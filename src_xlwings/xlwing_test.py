#! /usr/bin/python
# -*- coding: utf-8 -*-
import xlwings as xw
from xlwings import Range
import openpyxl
'''
app = xw.App(visible=False, add_book=False)
app.display_alerts = False
app.screen_updating = False
'''

wb_op = openpyxl.load_workbook(filename='1.xlsx',keep_vba=True)
ws_op = wb_op['Sheet1']
filllist = []
for i in ws_op.iter_rows(min_col=1, max_col=2, min_row=3,max_row=6 ):
    for cell in i:
        filllist.append(cell.fill)
wb_op.close()


'''
filepath = r'3.xlsm'
# filepath2 = r'2.xlsx'
wb = app.books.open(filepath)
# wb1 = app.books.open(filepath2)

ws = wb.sheets['DataMap']
for i in  range(20):
    ws.api.Rows(7+i).Insert()


wb.save(filepath)
wb.close()

'''
wb_op2 = openpyxl.load_workbook(filename=r'E:\java\python\Test\src_xlwings\a.xlsm', keep_vba=True)
ws_op2 = wb_op2['DataMap']
'''
for row in ws_op2.iter_rows(min_col=1, max_col=2, min_row=3,max_row=6 ):
    print(row)
    for cell in row:
        pass
wb_op2.save('3.xlsm')
wb_op2.close()
'''



