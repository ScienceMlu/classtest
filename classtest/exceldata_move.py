#! /usr/bin/python
# -*- coding: utf-8 -*-
import os
import time
import threading

from openpyxl import load_workbook, Workbook


class Deal(object):
    def __init__(self, dom_path):
        self.dom_path = dom_path
        i = 0
        if not os.path.exists('end%d' % i):
            os.mkdir('end%d' % i)
            out_path = 'end%d' % i
        else:
            i += 1
            os.mkdir('end%d' % i)
            out_path = 'end%d' % i
        self.out_path = r'%s\%s' % (dom_path, out_path)

    def find_files(self):
        data_excel = []
        old_data_path = r'%s\old' % self.dom_path
        os.chdir(old_data_path)
        files = os.listdir(old_data_path)
        for F in files:
            if os.path.splitext(F)[-1] == '.xlsx':
                data_excel.append(F)
        return data_excel


class BaseUse(object):
    def __init__(self, filename, sheetname):
        keyword_old = ['啊', '是', '的', '风', '个', '好', '就', '看', '了', '在']
        keyword_new = ['I', 'II', 'III', 'IV', 'V', 'VI']
        self.keyword_old = keyword_old
        self.keyword_new = keyword_new
        self.filename = filename
        self.wb = load_workbook(filename)
        self.ws = self.wb[sheetname]

    def cols_catch_row(self, column, keyword_us):
        key_use_row = []
        for cell in self.ws.cell(column):
            if cell.value == keyword_us:
                key_use_row.append(cell.row)
        return key_use_row

    def keyRow_catch(self, keyword):
        keyRow = None
        keyCol = None
        for u_row in self.ws.iter_rows():
            for cell in u_row:
                if cell.value == keyword:
                    keyRow = cell.row
        return keyRow, keyCol

    def old_data_catch(self, keyword):
        data_list = []
        row_use = self.keyRow_catch(keyword)[0]
        Nomax = self.ws.max_row - row_use
        for c_row in self.ws.iter_rows():
            for cell in c_row:
                if cell.value == keyword:
                    for No in range(Nomax):
                        data_list.append(self.ws['%s%d' % (cell.column, cell.row + No + 1)].value)
        return data_list

    def add_special_data(self, keyword, count, special_dict=dict, out_data_list=list):
        out_data_list.insert(index=count, object=special_dict[keyword])
        return out_data_list

    def mix_2(self, data_list_base, data_list_use, mix_format):
        mix_data = []
        for data_key, data_value in enumerate(data_list_base):
            mix_data.append(mix_format % (data_value, data_list_use[data_key]))
        print(mix_data)
        return mix_data

    def mix_3(self, data_list_base, data_list_use1, data_list_use2, mix_format):
        mix_data = []
        for data_key, data_value in enumerate(data_list_base):
            mix_data.append(mix_format % (data_value, data_list_use1[data_key], data_list_use2[data_key]))
        print(mix_data)
        return mix_data

    def data_save(self):
        old_multi_data = []
        first_D = self.old_data_catch(keyword=self.keyword_old[0])
        old_multi_data.append(first_D)
        second_D = self.old_data_catch(keyword=self.keyword_old[1])
        old_multi_data.append(second_D)
        third_D = self.old_data_catch(keyword=self.keyword_old[2])
        old_multi_data.append(third_D)
        forth_D = self.old_data_catch(keyword=self.keyword_old[3])
        old_multi_data.append(forth_D)
        fifth_D = self.old_data_catch(keyword=self.keyword_old[4])
        old_multi_data.append(fifth_D)
        sixth_D = self.old_data_catch(keyword=self.keyword_old[5])
        old_multi_data.append(sixth_D)
        seventh_D = self.old_data_catch(keyword=self.keyword_old[6])
        old_multi_data.append(seventh_D)
        eight_D = self.old_data_catch(keyword=self.keyword_old[7])
        old_multi_data.append(eight_D)
        ninth_D = self.old_data_catch(keyword=self.keyword_old[8])
        old_multi_data.append(ninth_D)
        ten_D = self.old_data_catch(keyword=self.keyword_old[9])
        old_multi_data.append(ten_D)
        return old_multi_data
        # =========


class DataWrite(object):
    def __init__(self, deal_path, old_name, keyword, data):
        os.chdir(deal_path)
        self.keyword = keyword
        self.data = data
        wb = Workbook()
        ws = wb.create_sheet('output', index=1)
        self.wb = wb
        self.ws = ws
        self.filename = 'NEW_%s' % old_name
        wb.save(self.filename)

    def cell_format(self):
        for row in self.ws.max_row:
            self.ws.row_dimensions[row].height = 83

    def write_data(self):
        Nomax = len(self.data)
        for row in self.ws.iter_rows():
            for cell in row:
                if cell.value == self.keyword:
                    for No in range(Nomax):
                        self.ws['%s%d' % (cell.column, cell.row + No + 1)].value = self.data[No]
        self.wb.save(self.filename)


# 运行函数


def newfile_format(filename_new):
    keyword_new = ['I', 'II', 'III', 'IV', 'V', 'VI']
    wb = load_workbook(filename=filename_new)
    ws = wb['output']
    dict_cols = {0: 'A', 1: 'B', 2: 'C', 3: 'D', 4: 'E', 5: 'F', 6: 'G'}
    for key, key_value in enumerate(keyword_new):
        ws['%s%d' % (dict_cols[key], key)].value = key_value


def use_get(file):
    old_deal = BaseUse(filename=file, sheetname='outlook')
    old_data_list1 = old_deal.data_save()
    old_data_mix2 = old_deal.mix_2(old_data_list1[0], old_data_list1[1], mix_format=r'[%s]%s')
    old_data_mix2_1 = old_deal.mix_2(old_data_list1[3], old_data_list1[4], mix_format=r'%s：%s')
    old_data_mix3 = old_deal.mix_3(old_data_list1[7], old_data_list1[8], old_data_list1[9], mix_format=r'%s/%s/%s')
    return old_deal, old_data_list1, old_data_mix2, old_data_mix2_1, old_data_mix3


def special_data_get(file, No):  # 输出字典{因子：水准:'', 确认项目：， 期待值：}
    global que_row
    special_data = BaseUse(filename=file, sheetname='specical')
    keyword_use = ['因子:水准', '確認', '期待值']
    cols_No = special_data.keyRow_catch(keyword=No)[1]  # No的圈值所在列
    cols_yin = special_data.keyRow_catch(keyword='因子')[1]
    cols_shui = special_data.keyRow_catch(keyword='水准')[1]
    cols_que = special_data.keyRow_catch(keyword='確認')[1]
    cols_qi = special_data.keyRow_catch(keyword='期待值')[1]
    special_row = special_data.cols_catch_row(column=cols_No, keyword_us='〇')  # 圈值的行数
    que_row = special_data.cols_catch_row(column=cols_No, keyword_us='●')  # 黑圈的行数
    # 因子datalist
    cols_yin_data = []
    for row in special_row:
        cols_yin_data.append(special_data.ws['%s%d' % (cols_yin, row)])
    # 水准datalist
    cols_shui_data = []
    for row_s in special_row:
        cols_shui_data.append(special_data.ws['%s%d' % (cols_shui, row_s)])
    # mix因子：水准
    special_out = special_data.mix_2(data_list_base=cols_yin_data, data_list_use=cols_shui_data, mix_format='%s：%s')
    # 确认项目和期待值
    que_value = special_data.ws['%s%d' % (cols_que, que_row[0])]
    qi_value = special_data.ws['%s%d' % (cols_qi, que_row[0])]
    # 创建输出字典
    out_data = dict.fromkeys(keyword_use)
    out_data[keyword_use[0]] = special_out
    out_data[keyword_use[1]] = que_value
    out_data[keyword_use[2]] = qi_value

    return out_data


def use_write(file, c_data, Deal_name):
    D1 = DataWrite(deal_path=Deal_name.out_path, old_name=file, keyword=c_data[0].keyword_new[0], data=c_data[2])
    newfile_format(filename_new=F)
    D1.write_data()
    D2 = DataWrite(deal_path=Deal_name.out_path, old_name=file, keyword=c_data[0].keyword_new[1], data=c_data[1][2])
    D2.write_data()
    D3 = DataWrite(deal_path=Deal_name.out_path, old_name=file, keyword=c_data[0].keyword_new[2], data=c_data[3])
    D3.write_data()
    D4 = DataWrite(deal_path=Deal_name.out_path, old_name=file, keyword=c_data[0].keyword_new[3], data=c_data[4])
    D4.write_data()
    D5 = DataWrite(deal_path=Deal_name.out_path, old_name=file, keyword=c_data[0].keyword_new[4], data=c_data[1][5])
    D5.write_data()
    D6 = DataWrite(deal_path=Deal_name.out_path, old_name=file, keyword=c_data[0].keyword_new[5], data=c_data[1][6])
    D6.write_data()


def get_thread_job(flies):
    global data_c
    print('get start\n')
    time.sleep(0.4)
    data_c = use_get(file=flies)
    time.sleep(0.4)
    print('T1 finish\n')
    # print('This is an added Thread ,number is %s' % threading.current_thread())


def write_thread_job(files, data, deal):
    print('write start\n')
    time.sleep(0.4)
    use_write(file=files, c_data=data, Deal_name=deal)
    time.sleep(0.4)
    print('write finish\n')


def main():
    t = Deal(dom_path=input('输入处理文件夹路径：'))
    file_list = t.find_files()
    for F in file_list:
        get_job = threading.Thread(target=get_thread_job(flies=F), name='get')
        write_job = threading.Thread(target=write_thread_job(files=F,data=data_c, deal=t), name='write')
        get_job.start()
        get_job.join(timeout=1)
        write_job.start()
        write_job.join(timeout=1)


if __name__ == '__main__':
    '''
    tool = BaseUse(filename='old.xlsx', keyword='tool')
    tool_data = tool.old_data_catch()
    series = BaseUse(filename='old.xlsx', keyword='series')
    series_data = series.old_data_catch()
    mix_base = EDataMix(filename='old.xlsx', mix_format=r'%d/%s')
    mix_data = mix_base.mix_2(tool_data, series_data)
    deal = DataWrite(filename='new.xlsx', keyword='mix', mix_data=mix_data)
    deal2 = DataWrite(filename='new.xlsx', keyword='tool', mix_data=tool_data)
    deal2.write_data()
    deal3 = DataWrite(filename='new.xlsx', keyword='series', mix_data=series_data)
    deal3.write_data()
    print(tool_data)
    print(series_data)
    print(mix_data)
    '''
    main()