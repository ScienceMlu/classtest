#! /usr/bin/python
# -*- coding: utf-8 -*-
import openpyxl
'''
filename=input('输入检查文件名：')
wb=openpyxl.load_workbook(filename)
'''
# 检查sheet内容


def check_tc(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet_name = wb.sheetnames
    group_data = []
    for j in range(len(sheet_name)):
        sheet1 = wb[sheet_name[j]]
        listA = []
        listB = []
        listC = []
        listD = []
        listE = []
        areaA = []
        areaB = []
        areaC = []
        areaD = []
        areaE = []
        countNumber = []
        for row in sheet1.iter_rows():
            for cell in row:
                if str(cell.fill.fgColor.rgb) == 'FFFFFF00':
                    if cell.value is None:
                        data_a = [cell.row, cell.column]
                        areaA.append(cell.coordinate)
                        listA.append(data_a)
                    elif cell.value == '〇':
                        print('ok')
                    elif str(cell.value).find('×') != -1:
                        data_b = [cell.row, cell.column]
                        areaB.append(cell.coordinate)
                        listB.append(data_b)
                    elif cell.value == '△':
                        data_c = [cell.row, cell.column]
                        areaC.append(cell.coordinate)
                        listC.append(data_c)
                    elif str(cell.value).find('N/A') != -1:
                        data_d = [cell.row, cell.column]
                        listD.append(data_d)
                        areaD.append(cell.coordinate)
                    else:
                        data_e = [cell.row, cell.column]
                        listE.append(data_e)
                        areaE.append(cell.coordinate)
        countNumber.append(len(areaA))
        countNumber.append(len(areaB))
        countNumber.append(len(areaC))
        countNumber.append(len(areaD))
        countNumber.append(len(areaE))
        group_data.append([areaA, areaB, areaC, areaD, areaE, countNumber])
        # group= dict.fromkeys(sheetName_now,[listA,listB,listC,listD,listE,areaA,areaB,areaC,areaD,areaE,countNumber])
        '''
        print(areaA)
        print(areaB)
        print(areaC)
        print(areaD)
        print(areaE)
        print(countNumber)
        print('------------')
        '''
    return group_data, sheet_name


def check_timetable(file_path):
    import re
    wb = openpyxl.load_workbook(file_path)
    sheet_name = wb.sheetnames
    error_txt = []
    for j in sheet_name:
        if j == '時間＿承認':
            sheet1 = wb[j]
            '''此处第一版，固定模式'''
            if sheet1['C9'].value is None:
                data_error_txt = '请填写检查时间（使用Ctrl+；）'
                error_txt.append(data_error_txt)
            if not re.match(u'^\u672a\u67e5\u8be2\u5230\u7ed3\u679c', sheet1['C13'].value):
                name_error_txt = '请填写检查人姓名'
                error_txt.append(name_error_txt)
            if sheet1['C15'].value is None:
                version_error_txt = '请输入版本号'
                error_txt.append(version_error_txt)
            if sheet1['C17'].value is None:
                pcname_error_txt = '请输入测试pc名'
                error_txt.append(pcname_error_txt)
            if sheet1['C19'].value is None:
                osAndlang_error_txt = '请输入pc的os名和言语环境'
                error_txt.append(osAndlang_error_txt)
            if not re.match(r'WLAN' | 'Bluetooth' | 'USB', sheet1['C21']):
                IF_error_txt = '请输入连接方式'
                error_txt.append(IF_error_txt)
            if sheet1['C23'].value is None:
                Model_error_txt = '请输入评价设备名'
                error_txt.append(Model_error_txt)
            if sheet1['C26'].value is None and not re.match('^\d', sheet1['C26'].value):
                time_error_txt = '请输入评价时间'
                error_txt.append(time_error_txt)
    return error_txt


# 写结果
def write_result(group_data, sheet_name, save_name, error_txt):
    from openpyxl.styles import Alignment
    from openpyxl import Workbook
    wb = Workbook()

    for k in range(len(sheet_name)):
        ws = wb.create_sheet(index=k)

        '''设置字体大小颜色，单元格背景'''

        # 合并单元格
        # 空值
        ws.merge_cells('A1:B2')
        # △
        ws.merge_cells('C1:D2')
        # ×
        ws.merge_cells('E1:F2')
        # N/A
        ws.merge_cells('G1:H2')
        # 格式不符
        ws.merge_cells('I1:J2')
        # 时间_承认 错误 标题
        ws.merge_cells('L1:O2')
        # 检查日错误
        ws.merge_cells('L3:O4')
        # 检查人错误
        ws.merge_cells('L5:O6')
        # 版本号错误
        ws.merge_cells('L7:O8')
        # 测试pc名错误
        ws.merge_cells('L9:O10')
        # 测试pc的os和言语环境错误
        ws.merge_cells('L11:O12')
        # 连接方式错误
        ws.merge_cells('L13:O14')
        # 评价设备名错误
        ws.merge_cells('L15:O16')
        # 评价时间错误
        ws.merge_cells('L17:O18')

        # 居中单元格
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['C1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['E1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['G1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['I1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['L1'].alignment = Alignment(horizontal='center', vertical='center')
        ws['L3'].alignment = Alignment(horizontal='center', vertical='center')
        ws['L5'].alignment = Alignment(horizontal='center', vertical='center')
        ws['L7'].alignment = Alignment(horizontal='center', vertical='center')
        ws['L9'].alignment = Alignment(horizontal='center', vertical='center')
        ws['L11'].alignment = Alignment(horizontal='center', vertical='center')
        ws['L13'].alignment = Alignment(horizontal='center', vertical='center')
        ws['L15'].alignment = Alignment(horizontal='center', vertical='center')
        ws['L17'].alignment = Alignment(horizontal='center', vertical='center')

        # 填写格式
        ws['A1'].value = '空值'
        ws['C1'].value = '△'
        ws['E1'].value = '×'
        ws['G1'].value = 'N/A'
        ws['I1'].value = '格式不符'
        ws['L1'].value = '時間＿承認 错误'

        i = 0
        for cols in range(1, 10, 2):
            ws.cell(row=3, column=cols, value='个数')
        for cols in range(2, 12, 2):
            ws.cell(row=3, column=cols, value='位置')

        # 写入位置数据
        # 空值位置
        data_len = len(group_data[k][0]) + 4
        print(data_len - 4)
        for rows in range(4, data_len):
            ws.cell(column=2, row=rows, value=group_data[k][0][i])
            if i > data_len - 4:
                break
            i = i + 1
        i = 0

        # △位置

        data_len = len(group_data[k][1]) + 4
        print(data_len - 4)
        for rows in range(4, data_len):
            ws.cell(column=4, row=rows, value=group_data[k][1][i])
            if i > data_len - 4:
                break
            i = i + 1
        i = 0

        # ×位置

        data_len = len(group_data[k][2]) + 4
        print(data_len - 4)
        for rows in range(4, data_len):
            ws.cell(column=6, row=rows, value=group_data[k][2][i])
            if i > data_len - 4:
                break
            i = i + 1
        i = 0

        # N/A位置

        data_len = len(group_data[k][3]) + 4
        print(data_len - 4)
        for rows in range(4, data_len):
            ws.cell(column=8, row=rows, value=group_data[k][3][i])
            if i > data_len - 4:
                break
            i = i + 1
        i = 0

        # 格式不符合位置

        data_len = len(group_data[k][4]) + 4
        print(data_len - 4)
        for rows in range(4, data_len):
            ws.cell(column=10, row=rows, value=group_data[k][4][i])
            if i > data_len - 4:
                break
            i = i + 1
        i = 0
        # 个数
        data_len = len(group_data[k][5])
        print(data_len)
        for cols in range(1, 10, 2):
            ws.cell(row=4, column=cols, value=group_data[k][5][i])
            # areaA:0,areaB:1,areaC:2,areaD:3,areaE:4,countNumber:5
            i += 1
            if i > data_len:
                break
        print('================')

        # 时间_承认错误文本
        for j in error_txt:
            for m in range(1, 17, 2):
                cell_name = 'L%d' % m
                ws[cell_name].value = j
        # 保存
    wb.save(u'%s' % save_name)


'''                
print('Wait a moment.....')



t=open('coordinate.txt','w')
if(listA!=[]):
    t.write('空值='+pprint.pformat(listA) + '\n' +'个数='+str(len(listA))+'\n'+'========================='+'\n')
if(listB!=[]):
    t.write('×所在line=' + pprint.pformat(listB)  + '\n' +'个数='+str(len(listB))+ '\n' + '====================='+'\n')
if(listC!=[]):
    t.write('△所在line=' + pprint.pformat(listC)  + '\n' +'个数='+str(len(listC))+ '\n' + '==================='+'\n')
if(listD!=[]):
    t.write('N/A所在line='+pprint.pformat(listD)  + '\n' +'个数='+str(len(listD))+ '\n' + '====================='+'\n')
if(listE!=[]):
    t.write('书写不规范所在line=' + pprint.pformat(listE) + '\n' +'个数='+str(len(listE))+ '\n'+'==============='+'\n')
t.close()
print('OK,Thank you for using!!!!')
'''
