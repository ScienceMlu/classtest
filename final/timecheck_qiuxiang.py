#!/usr/bin/env python
# -*- coding: utf-8 -*-
# -*- coding:utf-8 -*-
import openpyxl
import os
import time
from datetime import datetime
import openpyxl.styles

# 关键字初始化
from openpyxl.cell import cell

wb = openpyxl.load_workbook(filename=u'1.xlsx')
ws = wb['時間_承認']

dict_key = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
dict_use = dict.fromkeys(dict_key)
for i_key, i_value in enumerate(dict_key):
    dict_use[i_value] = i_key + 1
print(list(dict_use.keys())[list(dict_use.values()).index(5)])
date_list = []


def checkline(message):
    global lastCell, currentCell
    lastCell = None
    for index in range(13):
        if currentCell.colunm not in dict_use:
            continue
        lastCol = dict_use.get(currentCell.colunm) + 1
        lastCell = ws.cell(currentCell.row, lastCol)
        if currentCell.value is not None:
            date_list.append(currentCell.value)
        elif currentCell.value is None and lastCell.value is not None:
            print(message, currentCell)
        currentCell = lastCell
        if currentCell.value is not None and not is_valid_time(currentCell.value):
            print('检查日格式有误：', currentCell)
    return date_list


def is_valid_time(dataStr):
    try:
        strp = datetime.strptime(str(dataStr), '%Y-%m-%d %H:%M:%S')
        return True
    except:
        return False


for row in ws.iter_rows():
    for currentCell in row:
        if currentCell.value == '检查日':
            date_list = checkline('检查日有误：')

list1 = []
currentCell1 = []
for row1 in ws.iter_rows():
    for currentCell1 in row1:
        if currentCell1.value == str('检查者') and currentCell1.fill != openpyxl.styles.fills.GradientFill(
                stop=['FFFF00FF', 'FFFF00FF']):
            if currentCell1.value is not None:
                list1.append(currentCell1.value)

list2 = []
for row2 in ws.iter_rows():
    for currentCell2 in row2:
        if currentCell2.value == str('Version') and not currentCell2.fill.bgColor.rgb == 'FFFF00FF':
            if currentCell2.value is not None:
                list2.append(currentCell2.value)
        if date_list is not None and list2 is None:
            print('Version漏记：', currentCell2)

list3 = []
for row3 in ws.iter_rows():
    for currentCell3 in row3:
        if currentCell3.value == str('PC名'):
            if currentCell3.value is not None:
                list3.append(currentCell3.value)
        if date_list is not None and list3 is None:
            print('PC名漏记：', currentCell3)

list4 = []
for row4 in ws.iter_rows():
    for currentCell4 in row4:
        if currentCell4.value == str('OSの種類/言語'):
            if currentCell4.value is not None:
                list4.append(currentCell4.value)
        if date_list is not None and list4 is None:
            print('OSの種類/言語漏记：', currentCell4)

list5 = []
for row5 in ws.iter_rows():
    for currentCell5 in row5:
        if currentCell5.value == str('I/F漏记'):
            if currentCell5.value is not None and not currentCell5.value.re.search(r'USB').re.search(r'LAN').re.search(
                    r'WLAN'):
                list4.append(currentCell5.value)
        if date_list is not None and list5 is None:
            print('I/F漏记：', currentCell5)

list6 = []
for row6 in ws.iter_rows():
    for currentCell6 in row4:
        if currentCell6.value == str('評価時間'):
            if currentCell6.value is not None:
                list4.append(currentCell6.value)
        if date_list is not None and list6 is None:
            print('評価時間漏记：', currentCell6)
            if currentCell6.value is not None and currentCell6.value != time.strftime('%.2f', time.localtime()):
                print('評価時間格式有误：', currentCell6)


def checkmode2(ws, keyword, errortxt):
    list_d = []
    for row in ws.iter_rows():
        for currentCell in row:
            if currentCell.value == str(keyword):
                if currentCell.value is not None:
                    list_d.append(currentCell.value)
            if date_list is not None and list_d is None:
                print('%s：%s' % (errortxt, currentCell))
