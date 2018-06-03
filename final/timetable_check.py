#!/usr/bin/env python
# -*- coding: utf-8 -*-
# -*- coding:utf-8 -*-
import openpyxl
import os
import time
from datetime import datetime
import openpyxl.styles

# 关键字初始化
from openpyxl.cell import cell

dict_key = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
dict_use = dict.fromkeys(dict_key)
for i_key, i_value in enumerate(dict_key):
    dict_use[i_value] = i_key + 1
print(list(dict_use.keys())[list(dict_use.values()).index(5)])

keyword_list = ['检查日', '检查者', 'Version', 'PC名', 'OSの種類/言語', 'I/F', '評価Model', '評価時間']

errortxt_list = ['检查日漏记：', '检查者漏记：', '版本漏记：', 'PC名漏记：', 'OSの種類/言語漏记：',
                 'I/F漏记：', '評価Model漏记：', '評価時間漏记：']
error_dict = dict.fromkeys(errortxt_list)
user_list = []
date_list = []
ver_list = []
pc_list = []
os_list = []
IF_list = []
model_list = []
checktime_list = []
data_tup = (date_list, user_list, ver_list, pc_list, os_list, IF_list, model_list, checktime_list)
for d_key, d_value in enumerate(errortxt_list):
    error_dict[d_value] = data_tup[d_key]
"""
def checkline(message):
    global lastCell, currentCell
    lastCell = None
    for index in range(13):
        if currentCell.colunm not in dict_use:
            continue
        lastCol = dict_use.get(currentCell.colunm) + 1
        lastCell = ws.cell(currentCell.row, lastCol)
        if currentCell.value is not None:
            date_list.append(currentCell.value)
        elif currentCell.value is None and lastCell.value is not None:
            print(message, currentCell)
        currentCell = lastCell
        if currentCell.value is not None and not is_valid_time(currentCell.value):
            print('检查日格式有误：', currentCell)
    return date_list


def is_valid_time(dataStr):
    try:
        strp = datetime.strptime(str(dataStr), '%Y-%m-%d %H:%M:%S')
        return True
    except:
        return False
"""

'''
for row in ws.iter_rows():
    for currentCell in row:
        if currentCell.value == '检查日':
            date_list = checkline('检查日有误：')
'''

def value_cell(tcell):
    t_coordinate = '%s%d' % (list(dict_use.keys())[list(dict_use.values()).index(dict_use[tcell.column] + 1)], tcell.row)
    return t_coordinate


def checkmode_value_exist(ws,key_list=list, errortxt_list=list, l_key=int):
    for row in ws.iter_rows():
        for t_cell in row:
            if t_cell.value == '%s' % key_list[l_key] and ws[value_cell(t_cell)].value is None:
                error_dict[errortxt_list[l_key]].append(ws[value_cell(t_cell)].coordinate)
    print(error_dict)

def checkmode2(ws, keyword, errortxt):
    list_d = []
    for row in ws.iter_rows():
        for currentCell in row:
            if currentCell.value == str(keyword):
                if currentCell.value is not None:
                    list_d.append(currentCell.value)
            if date_list is not None and list_d is None:
                print('%s：%s' % (errortxt, currentCell))


wb = openpyxl.load_workbook(filename=u'1.xlsx')
ws = wb['時間_承認']


